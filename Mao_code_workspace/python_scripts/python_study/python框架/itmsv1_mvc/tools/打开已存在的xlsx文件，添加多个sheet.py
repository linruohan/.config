import pandas as pd 
import numpy as np
from openpyxl import load_workbook
"添加时，如果存在，02后面021,021,023一次添加"
path='001.xlsx'
book = load_workbook(path)
writer = pd.ExcelWriter(path,engine ='openpyxl')
writer.book = book
a = pd.DataFrame(np.random.random((3,2)))
a.to_excel(writer,'a1 ')
a.to_excel(writer,'a2')
writer.save()


"添加时，如果存在，02后面021,021,023一次添加"
path='001(pandas).xlsx'
book = load_workbook(path)
writer = pd.ExcelWriter(path,engine ='openpyxl')
writer.book = book

df=pd.read_excel(path,index = [ 0])
df.to_excel(writer,'sheet111',startrow = 1,startcol = 7)
df1=pd.read_excel(path,index = [1])
df1.to_excel(writer,'sheet123',startrow = 0,startcol = 0)
df2=pd.read_excel(path,index = [2])
df2.to_excel(writer,'sheet123',startrow = 3,startcol = 3)
writer.save()